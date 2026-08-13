from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert

from app_prontocardio.database import postgres_engine
from app_prontocardio.models import Tiss

LINHA_INICIAL = 7
LIMITE_LINHAS_VAZIAS = 50
NOME_PLANILHA = 'Tab 38'


def _argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            'Importa a Tabela 38 (mensagens de glosa) da planilha oficial '
            'de terminologias TISS publicada pela ANS.'
        )
    )
    origem = parser.add_mutually_exclusive_group(required=True)
    origem.add_argument(
        '--arquivo',
        type=Path,
        help='Planilha oficial de terminologias TISS da ANS.',
    )
    origem.add_argument(
        '--fhir-json',
        type=Path,
        help='CodeSystem JSON oficial da TUSS 38 publicado pela ANS.',
    )
    parser.add_argument('--versao', required=True)
    parser.add_argument('--aplicar', action='store_true')
    parser.add_argument(
        '--confirmar-gravacao',
        action='store_true',
        help='Confirma explicitamente a gravação na conexão configurada.',
    )
    return parser.parse_args()


def _normalizar_data(valor: object) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


def carregar_terminologia(arquivo: Path, versao: str) -> list[dict]:
    if not arquivo.is_file():
        raise FileNotFoundError(f'Planilha não encontrada: {arquivo}')

    workbook = load_workbook(arquivo, read_only=True, data_only=True)
    try:
        if NOME_PLANILHA not in workbook.sheetnames:
            raise ValueError(
                f'A aba {NOME_PLANILHA!r} não existe na planilha.'
            )

        worksheet = workbook[NOME_PLANILHA]
        itens: list[dict] = []
        linhas_vazias = 0
        for valores in worksheet.iter_rows(
            min_row=LINHA_INICIAL,
            max_col=5,
            values_only=True,
        ):
            codigo, termo, inicio, fim, fim_implantacao = valores
            if codigo is None and termo is None:
                linhas_vazias += 1
                if linhas_vazias >= LIMITE_LINHAS_VAZIAS:
                    break
                continue

            linhas_vazias = 0
            codigo_normalizado = str(codigo or '').strip()
            termo_normalizado = str(termo or '').strip()
            if not codigo_normalizado or not termo_normalizado:
                continue

            if codigo_normalizado.endswith('.0'):
                codigo_normalizado = codigo_normalizado[:-2]

            itens.append(
                {
                    'codigo_termo': codigo_normalizado,
                    'termo': termo_normalizado,
                    'dt_inicio_vigencia': _normalizar_data(inicio),
                    'dt_fim_vigencia': _normalizar_data(fim),
                    'dt_fim_implantacao': _normalizar_data(fim_implantacao),
                    'fonte': (
                        'ANS - Padrão TISS, Tabela 38, versão '
                        f'{versao}'
                    ),
                    'pagina_pdf': 0,
                }
            )
        return itens
    finally:
        workbook.close()


def _conceitos_fhir(conceitos: list[dict[str, Any]]):
    for conceito in conceitos:
        yield conceito
        yield from _conceitos_fhir(conceito.get('concept', []))


def carregar_terminologia_fhir(
    arquivo: Path,
    versao: str,
) -> list[dict]:
    if not arquivo.is_file():
        raise FileNotFoundError(f'CodeSystem não encontrado: {arquivo}')

    with arquivo.open(encoding='utf-8') as stream:
        code_system = json.load(stream)
    if code_system.get('resourceType') != 'CodeSystem':
        raise ValueError('O JSON informado não é um CodeSystem FHIR.')

    itens = []
    for conceito in _conceitos_fhir(code_system.get('concept', [])):
        codigo = str(conceito.get('code') or '').strip()
        termo = str(conceito.get('display') or '').strip()
        if not codigo or not termo:
            continue
        itens.append(
            {
                'codigo_termo': codigo,
                'termo': termo,
                'dt_inicio_vigencia': None,
                'dt_fim_vigencia': None,
                'dt_fim_implantacao': None,
                'fonte': (
                    'ANS - CodeSystem FHIR TUSS 38, versão '
                    f'{versao}'
                ),
                'pagina_pdf': 0,
            }
        )
    return itens


def importar_terminologia(itens: list[dict]) -> None:
    if postgres_engine is None:
        raise RuntimeError('DATABASE_URL não configurada.')
    if not itens:
        raise RuntimeError('A planilha não possui termos TISS importáveis.')

    tabela = Tiss.__table__
    comando = insert(tabela).values(itens)
    comando = comando.on_conflict_do_update(
        index_elements=[tabela.c.codigo_termo],
        set_={
            'termo': comando.excluded.termo,
            'dt_inicio_vigencia': comando.excluded.dt_inicio_vigencia,
            'dt_fim_vigencia': comando.excluded.dt_fim_vigencia,
            'dt_fim_implantacao': comando.excluded.dt_fim_implantacao,
            'fonte': comando.excluded.fonte,
            'pagina_pdf': comando.excluded.pagina_pdf,
        },
    )
    with postgres_engine.begin() as connection:
        connection.execute(comando)


def main() -> None:
    args = _argumentos()
    if args.fhir_json:
        itens = carregar_terminologia_fhir(args.fhir_json, args.versao)
    else:
        itens = carregar_terminologia(args.arquivo, args.versao)
    print(f'Termos TISS encontrados: {len(itens)}')

    if not args.aplicar:
        print('Simulação concluída; nenhuma alteração foi gravada.')
        return
    if not args.confirmar_gravacao:
        raise RuntimeError(
            'Use --confirmar-gravacao para autorizar a escrita no banco.'
        )

    importar_terminologia(itens)
    print(f'Termos TISS gravados/atualizados: {len(itens)}')


if __name__ == '__main__':
    main()
