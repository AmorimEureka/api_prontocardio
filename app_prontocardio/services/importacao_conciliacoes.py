from collections import defaultdict
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from unicodedata import combining, normalize

from openpyxl import load_workbook

CENTAVOS = Decimal('0.01')


class PlanilhaConciliacaoInvalidaError(ValueError):
    pass


@dataclass(frozen=True)
class LinhaConciliacaoPlanilha:
    numero_linha: int
    numero_nfse: str
    processo_recebimento: str
    cd_remessa: int
    valor_alocado: Decimal
    valor_glosado: Decimal
    data_previsao_recebimento: date | None

    @property
    def chave(self) -> tuple[str, str]:
        return self.numero_nfse, self.processo_recebimento


@dataclass(frozen=True)
class VinculoConciliacaoAtual:
    conciliacao_id: int
    vinculo_id: int
    nfse_row_hash: str
    numero_nfse: str
    processo_recebimento: str
    cnpj_convenio: str
    valor_nfse: Decimal
    cd_remessa: int
    valor_alocado: Decimal
    valor_glosado: Decimal
    data_previsao_recebimento: date
    usuario_id: int

    @property
    def chave(self) -> tuple[str, str]:
        return self.numero_nfse, self.processo_recebimento

    @property
    def chave_nfse(self) -> tuple[str, str]:
        return self.numero_nfse, self.cnpj_convenio


@dataclass(frozen=True)
class AjusteVinculoPlanilha:
    atual: VinculoConciliacaoAtual
    linha: LinhaConciliacaoPlanilha


@dataclass(frozen=True)
class NovoVinculoPlanilha:
    referencia: VinculoConciliacaoAtual
    linha: LinhaConciliacaoPlanilha


@dataclass(frozen=True)
class PlanoReprocessamentoConciliacoes:
    grupos_carga: int
    grupos_analisados: int
    grupos_sem_remessa_numerica: int
    grupos_repetidos: int
    vinculos_esperados: int
    vinculos_presentes: int
    ajustes: tuple[AjusteVinculoPlanilha, ...]
    novos: tuple[NovoVinculoPlanilha, ...]


def _texto_chave(valor) -> str:
    if valor is None:
        return ''
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _dinheiro(valor) -> Decimal:
    if valor in (None, ''):
        return Decimal('0.00')
    texto = str(valor).strip().replace('R$', '').replace(' ', '')
    if ',' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    try:
        return Decimal(texto).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise PlanilhaConciliacaoInvalidaError(
            f'Valor monetário inválido na planilha: {valor!r}.'
        ) from exc


def _codigo_remessa(valor) -> int | None:
    texto = _texto_chave(valor)
    try:
        codigo = int(Decimal(texto))
    except (InvalidOperation, ValueError):
        return None
    if codigo <= 0:
        return None
    return codigo


def _data(valor, numero_linha: int) -> date | None:
    if valor in (None, ''):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for formato in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass
    raise PlanilhaConciliacaoInvalidaError(
        f'Data de previsão inválida na linha {numero_linha}: {valor!r}.'
    )


def _normalizar_cabecalho(valor) -> str:
    texto = normalize('NFKD', str(valor or ''))
    sem_acentos = ''.join(
        caractere for caractere in texto if not combining(caractere)
    )
    return ' '.join(sem_acentos.upper().strip().split())


def _localizar_cabecalho(planilha) -> tuple[int, dict[str, int]]:
    obrigatorios = {
        'COLUNA1',
        'GLOSA',
        'VLR LIQ NF',
        'NF',
        'PREV. RECBTO.',
        'PROCESSO',
    }
    for numero_linha, valores in enumerate(
        planilha.iter_rows(min_row=1, max_row=20, values_only=True),
        start=1,
    ):
        indices = {
            _normalizar_cabecalho(valor): indice
            for indice, valor in enumerate(valores)
            if valor is not None
        }
        if obrigatorios.issubset(indices):
            return numero_linha, indices
    raise PlanilhaConciliacaoInvalidaError(
        'Cabeçalhos da aba BASE não foram encontrados.'
    )


def ler_linhas_conciliacao_planilha(
    caminho: str | Path,
    nome_aba: str = 'BASE',
) -> list[LinhaConciliacaoPlanilha]:
    arquivo = Path(caminho)
    if not arquivo.is_file():
        raise PlanilhaConciliacaoInvalidaError(
            f'Planilha não encontrada: {arquivo}.'
        )
    workbook = load_workbook(arquivo, read_only=True, data_only=True)
    if nome_aba not in workbook.sheetnames:
        raise PlanilhaConciliacaoInvalidaError(
            f'Aba {nome_aba!r} não encontrada na planilha.'
        )
    planilha = workbook[nome_aba]
    linha_cabecalho, indices = _localizar_cabecalho(planilha)
    linhas = []
    for numero_linha, valores in enumerate(
        planilha.iter_rows(
            min_row=linha_cabecalho + 1,
            values_only=True,
        ),
        start=linha_cabecalho + 1,
    ):
        numero_nfse = _texto_chave(valores[indices['NF']])
        processo = _texto_chave(valores[indices['PROCESSO']])
        remessa = valores[indices['COLUNA1']]
        valor_alocado = _dinheiro(valores[indices['VLR LIQ NF']])
        if not numero_nfse or not processo or remessa in (None, ''):
            continue
        if valor_alocado <= 0:
            continue
        cd_remessa = _codigo_remessa(remessa)
        if cd_remessa is None:
            continue
        linhas.append(
            LinhaConciliacaoPlanilha(
                numero_linha=numero_linha,
                numero_nfse=numero_nfse,
                processo_recebimento=processo,
                cd_remessa=cd_remessa,
                valor_alocado=valor_alocado,
                valor_glosado=_dinheiro(valores[indices['GLOSA']]),
                data_previsao_recebimento=_data(
                    valores[indices['PREV. RECBTO.']],
                    numero_linha,
                ),
            )
        )
    workbook.close()
    return linhas


def _validar_saldos_nfse(
    vinculos: list[VinculoConciliacaoAtual],
    ajustes: list[AjusteVinculoPlanilha],
    novos: list[NovoVinculoPlanilha],
) -> None:
    utilizados = defaultdict(lambda: Decimal('0.00'))
    limites = {}
    for vinculo in vinculos:
        utilizados[vinculo.chave_nfse] += vinculo.valor_alocado
        limites[vinculo.chave_nfse] = max(
            limites.get(vinculo.chave_nfse, Decimal('0.00')),
            vinculo.valor_nfse,
        )
    for ajuste in ajustes:
        chave = ajuste.atual.chave_nfse
        utilizados[chave] += (
            ajuste.linha.valor_alocado - ajuste.atual.valor_alocado
        )
    for novo in novos:
        utilizados[novo.referencia.chave_nfse] += novo.linha.valor_alocado

    excedentes = [
        (chave, valor, limites[chave])
        for chave, valor in utilizados.items()
        if valor > limites[chave]
    ]
    if excedentes:
        chave, utilizado, limite = sorted(excedentes)[0]
        raise PlanilhaConciliacaoInvalidaError(
            f'A NFS-e {chave[0]} excederia seu saldo: '
            f'R$ {utilizado} alocados para R$ {limite} disponíveis.'
        )


def _limitar_linhas_ao_saldo_nfse(
    chave: tuple[str, str],
    linhas: list[LinhaConciliacaoPlanilha],
    atuais: list[VinculoConciliacaoAtual],
    todos_vinculos: list[VinculoConciliacaoAtual],
) -> list[LinhaConciliacaoPlanilha]:
    referencia = min(atuais, key=lambda item: item.conciliacao_id)
    remessas_planilha = {linha.cd_remessa for linha in linhas}
    reservado = sum(
        (
            vinculo.valor_alocado
            for vinculo in todos_vinculos
            if vinculo.chave_nfse == referencia.chave_nfse
            and not (
                vinculo.chave == chave
                and vinculo.cd_remessa in remessas_planilha
            )
        ),
        Decimal('0.00'),
    )
    saldo = max(referencia.valor_nfse - reservado, Decimal('0.00'))
    limitadas = []
    for linha in sorted(
        linhas,
        key=lambda item: (item.numero_linha, item.cd_remessa),
    ):
        valor_alocado = min(linha.valor_alocado, saldo)
        if valor_alocado <= 0:
            raise PlanilhaConciliacaoInvalidaError(
                f'A NFS-e {linha.numero_nfse} não possui saldo para '
                f'vincular a remessa {linha.cd_remessa}.'
            )
        limitadas.append(replace(linha, valor_alocado=valor_alocado))
        saldo -= valor_alocado
    return limitadas


def planejar_reprocessamento_conciliacoes(
    linhas: list[LinhaConciliacaoPlanilha],
    vinculos: list[VinculoConciliacaoAtual],
) -> PlanoReprocessamentoConciliacoes:
    linhas_por_chave = defaultdict(list)
    for linha in linhas:
        linhas_por_chave[linha.chave].append(linha)

    vinculos_por_chave = defaultdict(list)
    for vinculo in vinculos:
        vinculos_por_chave[vinculo.chave].append(vinculo)

    ajustes = []
    novos = []
    grupos_analisados = 0
    grupos_repetidos = 0
    vinculos_esperados = 0
    vinculos_presentes = 0
    for chave, linhas_grupo in sorted(linhas_por_chave.items()):
        atuais = vinculos_por_chave.get(chave, [])
        if not atuais:
            continue
        grupos_analisados += 1
        vinculos_esperados += len(linhas_grupo)
        remessas_atuais = {item.cd_remessa for item in atuais}
        vinculos_presentes += sum(
            linha.cd_remessa in remessas_atuais for linha in linhas_grupo
        )
        if len(linhas_grupo) <= 1:
            continue
        grupos_repetidos += 1
        remessas = [linha.cd_remessa for linha in linhas_grupo]
        if len(remessas) != len(set(remessas)):
            raise PlanilhaConciliacaoInvalidaError(
                f'A NFS-e {chave[0]} e o processo {chave[1]} repetem '
                'a mesma remessa na planilha.'
            )
        linhas_limitadas = _limitar_linhas_ao_saldo_nfse(
            chave,
            linhas_grupo,
            atuais,
            vinculos,
        )
        atuais_por_remessa = {item.cd_remessa: item for item in atuais}
        referencia = min(atuais, key=lambda item: item.conciliacao_id)
        for linha in sorted(
            linhas_limitadas,
            key=lambda item: (item.numero_linha, item.cd_remessa),
        ):
            atual = atuais_por_remessa.get(linha.cd_remessa)
            if atual is None:
                novos.append(
                    NovoVinculoPlanilha(
                        referencia=referencia,
                        linha=linha,
                    )
                )
                continue
            if (
                atual.valor_alocado != linha.valor_alocado
                or atual.valor_glosado != linha.valor_glosado
            ):
                ajustes.append(
                    AjusteVinculoPlanilha(
                        atual=atual,
                        linha=linha,
                    )
                )

    _validar_saldos_nfse(vinculos, ajustes, novos)
    return PlanoReprocessamentoConciliacoes(
        grupos_carga=len(vinculos_por_chave),
        grupos_analisados=grupos_analisados,
        grupos_sem_remessa_numerica=(
            len(vinculos_por_chave) - grupos_analisados
        ),
        grupos_repetidos=grupos_repetidos,
        vinculos_esperados=vinculos_esperados,
        vinculos_presentes=vinculos_presentes,
        ajustes=tuple(ajustes),
        novos=tuple(novos),
    )
